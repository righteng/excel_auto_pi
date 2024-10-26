import openpyxl

#Excelを開く
wb = openpyxl.load_workbook('./サンプル.xlsx')

#Excelのシートを選択する
sheet = wb['Sheet1']

#Excelシートのセルに文字を入力する
sheet.cell(row=1, column=2).value = "値1"
#もしくは
#value = sheet['A1'].value

sheet.cell(row=1, column=3).value = "値2"

#Excelシートのセルの文字を出力する
ret = sheet.cell(row=1, column=2).value
print(ret, end = " ")

ret2 = sheet.cell(row=1, column=3).value
print(ret2)

#Excelシートのセルのデータを合計する
#合計をセルおよびプロンプトに出力する。
sum = 0
sum2 = 0
for row in range(2, 6):
    sheet.cell(row=row, column=2).value = row
    sheet.cell(row=row, column=3).value = 2 * row
    print(row, end = " ")
    print(2 * row)
    sum += sheet.cell(row=row, column=2).value
    sum2 += sheet.cell(row=row, column=3).value
sheet.cell(row=row+2, column=1).value = "合計"
sheet.cell(row=row+2, column=2).value = sum
sheet.cell(row=row+2, column=3).value = sum2
print("合計")
print(sum, end = " ")
print(sum2)
print("")

#グラフを作成する。
values = openpyxl.chart.Reference(sheet, min_col = 2, min_row = 2, max_col = 2, max_row = 5)
values2 = openpyxl.chart.Reference(sheet, min_col = 3, min_row = 2, max_col = 3, max_row = 5)
chart = openpyxl.chart.LineChart()
chart.add_data(values)
chart.add_data(values2)
sheet.add_chart(chart, 'A10')

#Excelを保存する
wb.save('./サンプル.xlsx')

#Excelを閉じる
wb.close()